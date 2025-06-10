import openpyxl
import numpy as np
import matplotlib.pyplot as plt


plt.style.use('bmh')

wb = openpyxl.load_workbook("normalized_workout_log.xlsx")

ws = wb.active


print('total number of rows: '+str(ws.max_row)+' . And total number of collumns: '+str(ws.max_column))

#get the size of ammount of exercises
exercises = ws.max_row - 1
print('Number of Exercises: ' + str(exercises))
# get the number of unique days exercised
allDays = []
for i in range(2, exercises):
    if ws.cell(row=i, column=1).value is not None:
        allDays.append(ws.cell(row=i, column=1).value)

TotalDays = [ws.cell(row=i,column=1).value for i in range(2, exercises + 2)]
print('Total Exercises: ' + str(len(TotalDays)) + '.')
TotalDaysDateOnly = [d.date() if hasattr(d, 'date') else d for d in TotalDays]

daysExercised = np.unique(TotalDaysDateOnly)

print('Unique days exercsed: ' + ', '.join(day.strftime('%d-%m-%Y') for day in daysExercised))

x = np.arange(-2, 8, .1)
y = .1 * x ** 3 - x ** 2 + 3 * x + 2

fig = plt.figure(dpi=100, figsize=(10, 20), tight_layout=True)
available = ['default'] + plt.style.available
for i, style in enumerate(available):
    with plt.style.context(style):
        ax = fig.add_subplot(10, 3, i + 1)
        ax.plot(x, y)
    ax.set_title(style)

